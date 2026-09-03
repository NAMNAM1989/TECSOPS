# -*- coding: utf-8 -*-
"""Convert H21_SCSC.xlsx → data/scsc-h21/*.json (seed cho kho SCSC)."""
from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(r"e:\MẪU LÁT XÓA\KHO NGOÀI\H21_SCSC.xlsx")
OUT = Path(__file__).resolve().parents[1] / "data" / "scsc-h21"

UOM_MAP = {
    "UNK": "BAG",
    "UNA": "BAG",
    "UNC": "PCE",
    "PCE": "PCE",
    "BAG": "BAG",
    "SET": "SET",
    "PR": "PR",
    "MTR": "MTR",
    "UNIT": "UNIT",
    "KGM": "KGM",
}


def norm_uom(v, fallback="PCE"):
    s = str(v or "").strip().upper()
    return UOM_MAP.get(s, s or fallback)


def norm_hs(v):
    if v is None:
        return ""
    if isinstance(v, float):
        s = str(int(v)) if v == int(v) else str(v)
    else:
        s = str(v).strip()
    return re.sub(r"[^0-9]", "", s)[:12]


def num(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        n = float(v)
        return n if n == n else default
    except (TypeError, ValueError):
        return default


def make_id(i, cat, desc, hs):
    raw = f"{i}|{cat}|{desc}|{hs}".encode("utf-8")
    return f"scsc-h21-{hashlib.sha1(raw).hexdigest()[:10]}"


def main():
    if not SRC.exists():
        print(f"Missing source: {SRC}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["DATA SC"]
    items = []
    for r in range(2, (ws.max_row or 1) + 1):
        desc = ws.cell(r, 2).value
        if not desc or not str(desc).strip():
            continue
        category = str(ws.cell(r, 1).value or "").strip()
        description = re.sub(r"\s+", " ", str(desc).strip())
        hs = norm_hs(ws.cell(r, 3).value)
        origin = str(ws.cell(r, 4).value or "VIETNAM").strip().upper() or "VIETNAM"
        qty1 = num(ws.cell(r, 5).value)
        uom1 = norm_uom(ws.cell(r, 6).value, "PCE")
        qty2 = num(ws.cell(r, 7).value)
        uom2 = norm_uom(ws.cell(r, 8).value, "KGM")
        unit_price = num(ws.cell(r, 9).value)
        amount = num(ws.cell(r, 10).value)
        unit_factor = num(ws.cell(r, 11).value)
        if not amount and qty1 and unit_price:
            amount = round(qty1 * unit_price, 4)
        items.append(
            {
                "id": make_id(r, category, description, hs),
                "category": category,
                "description": description,
                "hsCode": hs,
                "origin": origin,
                "qty1": qty1,
                "uom1": uom1,
                "qty2": qty2,
                "uom2": uom2,
                "unitPrice": round(unit_price, 6),
                "amount": round(amount, 4),
                "unitFactor": unit_factor,
                "sortOrder": len(items),
                "warehouseScope": "SCSC",
                "active": True,
            }
        )

    stamps = []
    ws2 = wb["STAMP_ID"]
    for r in range(2, (ws2.max_row or 1) + 1):
        name = ws2.cell(r, 2).value
        stamp = ws2.cell(r, 3).value
        if not name or not stamp:
            continue
        stamp_id = str(stamp).strip().upper()
        stamps.append(
            {
                "id": f"stamp-{stamp_id.lower()}",
                "shipperName": str(name).strip(),
                "stampId": stamp_id,
                "warehouseScope": "SCSC",
            }
        )

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "catalog.json").write_text(
        json.dumps(
            {"version": 1, "warehouseScope": "SCSC", "items": items},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (OUT / "stamp-ids.json").write_text(
        json.dumps(
            {"version": 1, "warehouseScope": "SCSC", "items": stamps},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Wrote {len(items)} catalog + {len(stamps)} stamps → {OUT}")


if __name__ == "__main__":
    main()
