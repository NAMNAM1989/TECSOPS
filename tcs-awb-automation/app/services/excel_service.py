from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.data.models import AwbJobResult
from app.services.awb_service import mark_duplicates, validate_row


HEADERS = [
    "AWB",
    "ACTION",
    "FLIGHT_DATE",
    "FLIGHT_NO",
    "PCS",
    "GROSS_WEIGHT",
    "DOCUMENT_TYPE",
    "PRINT_COPIES",
    "NOTE",
]


def create_import_template(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "AWB_IMPORT"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["123-12345678", "LOOKUP", "", "", "", "", "AWB", 1, "mẫu tra cứu"])
    ws.append(["12312345679", "PRINT", "", "", "", "", "AWB", 1, "mẫu in nếu hoàn thành"])
    wb.save(path)
    return path


def load_import_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "AWB_IMPORT" not in wb.sheetnames:
        raise ValueError("Thiếu sheet AWB_IMPORT")
    ws = wb["AWB_IMPORT"]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().upper() for c in next(rows_iter)]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if not any(raw):
            continue
        item = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
        out.append(item)
    return out


def excel_to_validated_rows(path: Path):
    raw_rows = load_import_excel(path)
    rows = [validate_row(r, i + 1) for i, r in enumerate(raw_rows)]
    return mark_duplicates(rows)


RESULT_HEADERS = [
    "STT",
    "AWB",
    "ACTION",
    "TCS_STATUS_RAW",
    "NORMALIZED_STATUS",
    "DOCUMENT_TYPE",
    "DOWNLOADED_FILE",
    "PRINT_STATUS",
    "START_TIME",
    "END_TIME",
    "DURATION_SECONDS",
    "RETRY_COUNT",
    "ERROR_CODE",
    "ERROR_MESSAGE",
]


_FILL_OK = PatternFill("solid", fgColor="C6EFCE")
_FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
_FILL_ERR = PatternFill("solid", fgColor="FFC7CE")


def _status_fill(status: str) -> PatternFill | None:
    s = (status or "").upper()
    if s in {"COMPLETED", "RECEPTION_COMPLETED", "DOWNLOADED", "PRINTED"}:
        return _FILL_OK
    if s in {"NOT_COMPLETED", "SKIPPED_DUPLICATE", "NEEDS_LOGIN", "PENDING"}:
        return _FILL_WARN
    if s in {"FAILED", "VALIDATION_ERROR", "SITE_CHANGED"}:
        return _FILL_ERR
    return None


def export_result_excel(results: list[AwbJobResult], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "RESULT"
    ws.append(RESULT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in results:
        row = [
            r.stt,
            r.awb,
            r.action,
            r.tcs_status_raw,
            r.normalized_status,
            r.document_type,
            r.downloaded_file,
            r.print_status,
            r.start_time,
            r.end_time,
            r.duration_seconds,
            r.retry_count,
            r.error_code,
            r.error_message,
        ]
        ws.append(row)
        fill = _status_fill(r.normalized_status)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    summary = {
        "Tổng AWB": len(results),
        "Thành công": sum(
            1
            for r in results
            if r.normalized_status in {"COMPLETED", "RECEPTION_COMPLETED", "DOWNLOADED", "PRINTED"}
        ),
        "Đã tiếp nhận": sum(1 for r in results if r.normalized_status == "RECEPTION_COMPLETED"),
        "Chưa hoàn thành": sum(1 for r in results if r.normalized_status == "NOT_COMPLETED"),
        "Lỗi dữ liệu": sum(1 for r in results if r.normalized_status == "VALIDATION_ERROR"),
        "Lỗi website": sum(1 for r in results if r.normalized_status in {"SITE_CHANGED", "FAILED"}),
        "Đã tải": sum(1 for r in results if r.downloaded_file),
        "Đã in": sum(1 for r in results if r.normalized_status == "PRINTED" or r.print_status == "OK"),
        "Bỏ qua trùng": sum(1 for r in results if r.normalized_status == "SKIPPED_DUPLICATE"),
    }
    ws2 = wb.create_sheet("SUMMARY")
    ws2.append(["Chỉ số", "Giá trị"])
    for k, v in summary.items():
        ws2.append([k, v])
    wb.save(path)
    return path