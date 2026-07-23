"""Excel mẫu + load cho khai báo ESID (tab KHAI BÁO ESID trên TCS).

TCS portal KHÔNG có nút upload Excel trên form — file này là:
1) Xuất từ Ops / chỉnh tay → agent Playwright điền form
2) Checklist field trước khi mở khóa REGISTER
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.services.esid_defaults import ESID_DEFAULT_PAYMENT_MODE
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Cột ổn định (agent parse). Thứ tự = thứ tự sheet ESID_DECLARE.
ESID_DECLARE_HEADERS: list[str] = [
    # Core
    "AWB",
    "FLIGHT_NO",
    "FLIGHT_DATE",  # YYYY-MM-DD hoặc DD-MM-YYYY
    "DEST",  # IATA 3
    "PCS",
    "GROSS_WEIGHT",  # tham chiếu Ops — form TCS hiện không có ô GW
    "TOTAL_HAWBS",
    "NATURE_OF_GOODS",
    "PAYMENT_MODE",  # mặc định Tiền mặt/Cash
    "CONSOL",  # 0|1
    "TECS_WAREHOUSE",  # 0|1 checkbox Kho TECS
    # Shipper
    "SHIPPER_NAME",
    "SHIPPER_ADDRESS",
    "SHIPPER_TEL",
    "SHIPPER_EMAIL",
    "SHIPPER_FAX",
    # Agent
    "AGENT_NAME",
    "AGENT_ADDRESS",
    "AGENT_TEL",
    "AGENT_EMAIL",
    "AGENT_FAX",
    "AGENT_VAT",
    # Consignee
    "CONSIGNEE_NAME",
    "CONSIGNEE_ADDRESS",
    "CONSIGNEE_TEL",
    "CONSIGNEE_EMAIL",
    "CONSIGNEE_FAX",
    "CONSIGNEE_VAT",
    # Notify
    "NOTIFY_NAME",
    "NOTIFY_ADDRESS",
    "NOTIFY_TEL",
    "NOTIFY_EMAIL",
    "NOTIFY_FAX",
    "NOTIFY_REMARK",
    # Handling (0|1)
    "SHC_PER",
    "SHC_PHARMA",
    "SHC_VAL",
    "SHC_AVI",
    "SHC_DGR",
    "SHC_BUP",
    "OTHER_REQUEST",
    # Người khai (bắt buộc nếu SUBMIT=1)
    "REGISTRANT_NAME",
    "REGISTRANT_TEL",
    "REGISTRANT_CCCD",
    # Control
    "SUBMIT",  # 0=dry-fill only, 1=cho phép HOÀN TẤT (cần confirm_register)
    "SHIPMENT_ID",
    "NOTE",
]

# Map cột Excel → id DOM TCS (live probe 2026-07-20)
EXCEL_TO_TCS_ID: dict[str, str] = {
    "AWB": "codAwbPfx+codAwbNum",
    "FLIGHT_NO": "flightNo",
    "FLIGHT_DATE": "datFltOri",
    "DEST": "codFds",
    "PCS": "qtyPcs",
    "TOTAL_HAWBS": "totalOfHawbs",
    "NATURE_OF_GOODS": "natureOfGoods",
    "PAYMENT_MODE": "codPayMod",
    "CONSOL": "shcConsol",
    "TECS_WAREHOUSE": "shcCod002",
    "SHIPPER_NAME": "shipperId",
    "SHIPPER_ADDRESS": "addressShp",
    "SHIPPER_TEL": "telShp",
    "SHIPPER_EMAIL": "emailShp",
    "SHIPPER_FAX": "faxShp",
    "AGENT_NAME": "agentId",
    "AGENT_ADDRESS": "addressAgt",
    "AGENT_TEL": "telAgt",
    "AGENT_EMAIL": "emailAgt",
    "AGENT_FAX": "faxAgt",
    "AGENT_VAT": "vatAgt",
    "CONSIGNEE_NAME": "consigneeId",
    "CONSIGNEE_ADDRESS": "addressCne",
    "CONSIGNEE_TEL": "telCne",
    "CONSIGNEE_EMAIL": "emailCne",
    "CONSIGNEE_FAX": "faxCne",
    "CONSIGNEE_VAT": "vatCne",
    "NOTIFY_NAME": "notifyId",
    "NOTIFY_ADDRESS": "addressNtf",
    "NOTIFY_TEL": "telNtf",
    "NOTIFY_EMAIL": "emailNtf",
    "NOTIFY_FAX": "faxNtf",
    "NOTIFY_REMARK": "desRmk001",
    "SHC_PER": "shcPer",
    "SHC_PHARMA": "shcOth002",
    "SHC_VAL": "shcVal",
    "SHC_AVI": "shcAvi",
    "SHC_DGR": "shcDgr",
    "SHC_BUP": "shcBup",
    "OTHER_REQUEST": "shcOthReq",
    "REGISTRANT_NAME": "shpRegNam",
    "REGISTRANT_TEL": "shpRegTel",
    "REGISTRANT_CCCD": "shpRegIdx",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_REQ_FILL = PatternFill("solid", fgColor="FFF2CC")
_MANUAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")

# Cột bắt buộc để dry-fill hữu ích
REQUIRED_FOR_FILL = {
    "AWB",
    "FLIGHT_NO",
    "DEST",
    "PCS",
    "SHIPPER_NAME",
    "SHIPPER_ADDRESS",
    "CONSIGNEE_NAME",
    "CONSIGNEE_ADDRESS",
}

# Cột Ops không có — phải nhập tay trước submit
MANUAL_BEFORE_SUBMIT = {"REGISTRANT_NAME", "REGISTRANT_TEL", "REGISTRANT_CCCD"}


def create_esid_declare_template(path: Path, *, with_sample: bool = True) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet dữ liệu ---
    ws = wb.active
    ws.title = "ESID_DECLARE"
    ws.append(ESID_DECLARE_HEADERS)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        h = ESID_DECLARE_HEADERS[i - 1]
        if h in REQUIRED_FOR_FILL:
            cell.fill = PatternFill("solid", fgColor="C65911")
        elif h in MANUAL_BEFORE_SUBMIT:
            cell.fill = _MANUAL_FILL
            cell.font = Font(bold=True)

    if with_sample:
        ws.append(_sample_row())

    for i, h in enumerate(ESID_DECLARE_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(22, len(h) + 2))
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ESID_DECLARE_HEADERS))}1"

    dv01 = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv01)
    for col_name in (
        "CONSOL",
        "TECS_WAREHOUSE",
        "SHC_PER",
        "SHC_PHARMA",
        "SHC_VAL",
        "SHC_AVI",
        "SHC_DGR",
        "SHC_BUP",
        "SUBMIT",
    ):
        idx = ESID_DECLARE_HEADERS.index(col_name) + 1
        letter = get_column_letter(idx)
        dv01.add(f"{letter}2:{letter}500")

    # --- Sheet hướng dẫn ---
    guide = wb.create_sheet("HUONG_DAN", 0)
    guide["A1"] = "Khai báo ESID nhanh — TECSOPS → Excel → TCS"
    guide["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "1) TCS portal KHÔNG có nút «Upload Excel» trên form KHAI BÁO ESID.",
        "   File này dùng để: xuất từ Ops → rà soát → agent Playwright điền form (hoặc copy tay).",
        "",
        "2) Luồng khuyến nghị (nhanh nhất):",
        "   Ops đã nhập lô → menu ⋮ «Excel khai báo ESID» → kiểm tra cột cam/cam nhạt",
        "   → bổ sung REGISTRANT_* (CCCD người khai) → agent dry-fill → người xác nhận → HOÀN TẤT.",
        "",
        "3) Màu cột sheet ESID_DECLARE:",
        "   Cam đậm = bắt buộc để điền form hữu ích (AWB, flight, dest, pcs, shipper/cnee).",
        "   Cam nhạt header = cột người khai (bắt buộc nếu SUBMIT=1).",
        "",
        "4) Combobox TCS (SHIPPER_NAME / AGENT_NAME / CONSIGNEE_NAME):",
        "   Form dùng Ant Select master — agent sẽ gõ tên rồi chọn option khớp.",
        "   Nếu TCS chưa có master → có thể chỉ điền được address/tel/email.",
        "",
        "5) FLIGHT_DATE: YYYY-MM-DD hoặc DD-MM-YYYY. Agent sẽ dùng nút CHỌN CHUYẾN BAY khi có.",
        "",
        "6) GROSS_WEIGHT: lấy từ Ops để đối chiếu — form TCS live hiện không có ô GW.",
        "",
        "7) SUBMIT=0 (mặc định): chỉ điền, KHÔNG tick đồng ý / KHÔNG bấm HOÀN TẤT.",
        "   SUBMIT=1: chỉ khi đã mở khóa REGISTER + confirm_register trên agent.",
        "",
        "8) Map cột → DOM TCS: xem sheet MAP_TCS.",
    ]
    for i, line in enumerate(lines, start=2):
        guide[f"A{i}"] = line
    guide.column_dimensions["A"].width = 110

    # --- Sheet map ---
    mp = wb.create_sheet("MAP_TCS")
    mp.append(["EXCEL_COLUMN", "TCS_DOM_ID", "NGUON_OPS", "GHI_CHU"])
    for c in mp[1]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    ops_hint = {
        "AWB": "shipment.awb",
        "FLIGHT_NO": "shipment.flight",
        "FLIGHT_DATE": "shipment.flightDate (+ sessionDate năm)",
        "DEST": "shipment.dest",
        "PCS": "shipment.pcs",
        "GROSS_WEIGHT": "shipment.kg (tham chiếu)",
        "NATURE_OF_GOODS": "goodsDescriptionPrint",
        "SHIPPER_NAME": "shipperNamePrint",
        "SHIPPER_ADDRESS": "shipperAddressPrint",
        "SHIPPER_TEL": "shipperPhonePrint",
        "SHIPPER_EMAIL": "shipperEmailPrint",
        "AGENT_NAME": "agentNamePrint",
        "AGENT_ADDRESS": "agentAddressPrint",
        "AGENT_TEL": "agentPhonePrint",
        "AGENT_EMAIL": "agentEmailPrint",
        "AGENT_VAT": "agentTaxCodePrint",
        "CONSIGNEE_NAME": "consigneeNamePrint",
        "CONSIGNEE_ADDRESS": "consigneeAddressPrint",
        "CONSIGNEE_TEL": "consigneePhonePrint",
        "CONSIGNEE_EMAIL": "consigneeEmailPrint",
        "CONSIGNEE_VAT": "taxCodePrint (nếu CNEE) / thủ công",
        "NOTIFY_NAME": "notifyNamePrint",
        "OTHER_REQUEST": "otherRequirementsPrint",
        "REGISTRANT_NAME": "(nhập tay)",
        "REGISTRANT_TEL": "(nhập tay)",
        "REGISTRANT_CCCD": "(nhập tay)",
        "SUBMIT": "mặc định 0",
        "SHIPMENT_ID": "shipment.id",
        "NOTE": "shipment.note",
    }
    for col in ESID_DECLARE_HEADERS:
        mp.append(
            [
                col,
                EXCEL_TO_TCS_ID.get(col, "—"),
                ops_hint.get(col, ""),
                "required" if col in REQUIRED_FOR_FILL else ("manual" if col in MANUAL_BEFORE_SUBMIT else ""),
            ]
        )
    for i, w in enumerate([22, 28, 40, 14], start=1):
        mp.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path


def _sample_row() -> list[Any]:
    row = {h: "" for h in ESID_DECLARE_HEADERS}
    row.update(
        {
            "AWB": "73807183061",
            "FLIGHT_NO": "VN0773",
            "FLIGHT_DATE": "2026-07-21",
            "DEST": "ICN",
            "PCS": 2,
            "GROSS_WEIGHT": 45.5,
            "NATURE_OF_GOODS": "GENERAL CARGO",
            "PAYMENT_MODE": ESID_DEFAULT_PAYMENT_MODE,
            "TOTAL_HAWBS": 0,
            "CONSOL": 0,
            "TECS_WAREHOUSE": 1,
            "SHIPPER_NAME": "CONG TY MAU TECS",
            "SHIPPER_ADDRESS": "123 Nguyen Van Linh, Q7, HCMC",
            "SHIPPER_TEL": "0901234567",
            "SHIPPER_EMAIL": "shipper@example.com",
            "AGENT_NAME": "TECS AGENT",
            "CONSIGNEE_NAME": "ABC TRADING CO",
            "CONSIGNEE_ADDRESS": "SEOUL, KOREA",
            "SUBMIT": 0,
            "NOTE": "mẫu — đổi AWB chưa khai trên TCS trước khi SUBMIT=1",
        }
    )
    return [row[h] for h in ESID_DECLARE_HEADERS]


def load_esid_declare_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "ESID_DECLARE" not in wb.sheetnames:
        raise ValueError("Thiếu sheet ESID_DECLARE")
    ws = wb["ESID_DECLARE"]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().upper() for c in next(rows_iter)]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if not any(raw):
            continue
        item = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
        out.append(item)
    return out


def analyze_row_readiness(row: dict[str, Any]) -> dict[str, Any]:
    missing_fill = [c for c in REQUIRED_FOR_FILL if not str(row.get(c) or "").strip()]
    missing_submit = [c for c in MANUAL_BEFORE_SUBMIT if not str(row.get(c) or "").strip()]
    submit = str(row.get("SUBMIT") or "0").strip() in {"1", "1.0", "true", "TRUE", "yes"}
    return {
        "awb": str(row.get("AWB") or ""),
        "can_dry_fill": len(missing_fill) == 0,
        "missing_for_fill": missing_fill,
        "can_submit": submit and len(missing_fill) == 0 and len(missing_submit) == 0,
        "missing_for_submit": missing_submit if submit else [],
        "submit_flag": submit,
        "warnings": [
            w
            for w in [
                "GROSS_WEIGHT có trong Excel nhưng form TCS hiện không có ô GW" if row.get("GROSS_WEIGHT") else "",
                "SUBMIT=1 nhưng REGISTER agent vẫn có thể bị khóa" if submit else "",
            ]
            if w
        ],
    }
