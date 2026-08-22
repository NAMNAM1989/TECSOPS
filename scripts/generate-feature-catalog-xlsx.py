#!/usr/bin/env python3
"""Sinh file Excel danh mục chức năng + điểm mạnh/yếu TECSOPS."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

OUT = Path(__file__).resolve().parents[1] / "docs" / "TECSOPS-Danh-muc-chuc-nang.xlsx"

# --- palette ---
NAVY = "1E3A5F"
NAVY2 = "0F2744"
TEAL = "0F766E"
WHITE = "FFFFFF"
SLATE = "334155"
MUTED = "64748B"
ROW_ALT = "F8FAFC"
GREEN = "166534"
GREEN_BG = "DCFCE7"
YELLOW = "854D0E"
YELLOW_BG = "FEF9C3"
ORANGE = "9A3412"
ORANGE_BG = "FFEDD5"
RED = "991B1B"
RED_BG = "FEE2E2"
BLUE_BG = "E0F2FE"
BLUE = "075985"
PURPLE_BG = "F3E8FF"
PURPLE = "6B21A8"
GOLD_BG = "FEF3C7"

THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color=NAVY)
FONT_H2 = Font(name="Calibri", size=13, bold=True, color=TEAL)
FONT_HEAD = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_CELL = Font(name="Calibri", size=10, color=SLATE)
FONT_BOLD = Font(name="Calibri", size=10, bold=True, color=SLATE)
FONT_SMALL = Font(name="Calibri", size=9, italic=True, color=MUTED)

WRAP = Alignment(wrap_text=True, vertical="center")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def style_header(ws, row: int, cols: int, color: str = NAVY) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = FONT_HEAD
        cell.fill = fill(color)
        cell.alignment = CENTER
        cell.border = THIN
    ws.auto_filter.ref = f"A{row}:{get_column_letter(cols)}{row}"
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = None  # set later with data range
    ws.freeze_panes = f"A{row + 1}"


def apply_row(ws, r: int, values: list, widths_align=None) -> None:
    alt = r % 2 == 0
    for c, val in enumerate(values, 1):
        cell = ws.cell(r, c, val)
        cell.font = FONT_CELL
        cell.alignment = WRAP
        cell.border = THIN
        if alt:
            cell.fill = fill(ROW_ALT)
    # rating column coloring (col 6 = Cong_dung if present)
    if len(values) >= 6 and isinstance(values[5], str):
        paint_rating(ws.cell(r, 6), values[5])
    if len(values) >= 7 and isinstance(values[6], (int, float)):
        paint_score(ws.cell(r, 7), values[6])


def paint_rating(cell, text: str) -> None:
    stars = text.count("★")
    cell.alignment = CENTER
    cell.font = Font(name="Calibri", size=10, bold=True, color=SLATE)
    if stars >= 5:
        cell.fill = fill(GREEN_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    elif stars == 4:
        cell.fill = fill(BLUE_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLUE)
    elif stars == 3:
        cell.fill = fill(YELLOW_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=YELLOW)
    elif stars == 2:
        cell.fill = fill(ORANGE_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
    else:
        cell.fill = fill(RED_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=RED)


def paint_score(cell, score) -> None:
    cell.alignment = CENTER
    try:
        n = float(score)
    except (TypeError, ValueError):
        return
    if n >= 5:
        cell.fill = fill(GREEN_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    elif n >= 4:
        cell.fill = fill(BLUE_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLUE)
    elif n >= 3:
        cell.fill = fill(YELLOW_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=YELLOW)
    elif n >= 2:
        cell.fill = fill(ORANGE_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
    else:
        cell.fill = fill(RED_BG)
        cell.font = Font(name="Calibri", size=10, bold=True, color=RED)


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def finish_filter(ws, header_row: int, last_row: int, cols: int) -> None:
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(cols)}{last_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_view.showGridLines = False


# =============================================================================
# DATA
# =============================================================================

FEATURES = [
    # id, nhom, ma, ten, viec_lam, ai_dung, cong_dung, diem, bat_buoc_ca, diem_manh, diem_yeu, ghi_chu
    ["A1", "A. Nền tảng", "AUTH", "Đăng nhập token / session", "Gate /api/auth, cookie HttpOnly 12h hoặc Bearer", "Mọi user", "★★★★", 4, "Có (prod)", "Chặn truy cập public Railway", "Token dùng chung, chưa phân role NV/admin", "Local có thể tắt unauthenticated"],
    ["A2", "A. Nền tảng", "SYNC", "Đồng bộ realtime đa máy", "Socket.IO + Postgres; mutation bus theo ngày phiên", "Phone + PC + VP", "★★★★★", 5, "Có", "Nhiều người cùng một bảng ngày", "Tự viết sync; dual blob+bảng phức tạp", "Trụ cột production"],
    ["A3", "A. Nền tảng", "OFFLINE", "Hàng đợi offline", "Queue mutation khi mất wifi, replay khi reconnect (≤500)", "NV kho", "★★★★★", 5, "Có", "Wifi kho yếu không mất DIM/lô", "Giới hạn 500; conflict last-write đơn giản", "Fail-before-apply khi đầy"],
    ["A4", "A. Nền tảng", "SESSION", "Ngày phiên (sessionDate)", "Mỗi ca một ngày local YYYY-MM-DD", "Tất cả", "★★★★★", 5, "Có", "Không lẫn dữ liệu hôm khác", "Phụ thuộc giờ máy local", "Sheet/portal/báo cáo đều bám ngày này"],
    ["A5", "A. Nền tảng", "WH", "4 kho tách biệt", "TECS-TCS, TECS-SCSC, TCS, SCSC + capability flags", "Tất cả", "★★★★★", 5, "Có", "Đúng nghiệp vụ 3 đội", "Dễ nhầm family vs OpsTeam khi báo cáo", "Sai kho = sai eSID/eCargo"],
    ["A6", "A. Nền tảng", "BADGE", "Badge sync Live / degraded / offline", "Hiển thị trạng thái kết nối", "NV kho", "★★★★", 4, "Không", "Tránh thao tác khi tưởng đã lưu", "User có thể bỏ qua badge", "UX cần thiết"],
    ["A7", "A. Nền tảng", "PG", "Postgres nguồn sự thật", "DATABASE_URL bắt buộc; không JSON runtime", "Hệ thống", "★★★★★", 5, "Có", "Redeploy Railway không mất ca", "Dual-write JSONB + bảng quan hệ nặng", "Backup/restore scripts có sẵn"],
    ["B1", "B. Bảng Ops", "DESKTOP", "Bảng lô desktop", "Lưới sửa nhanh AWB, chuyến, DEST, PCS, KG, KH, CNEE", "PC kho / VP", "★★★★★", 5, "Có", "Màn chính vận hành", "God-file AirCargoTracking (~1100 dòng)", "Hash route #/"],
    ["B2", "B. Bảng Ops", "MOBILE", "Thẻ lô + sheet sửa mobile", "Card + bottom sheet trên điện thoại", "NV cân", "★★★★★", 5, "Có", "Làm việc được khi không có PC", "Modal/sheet lớn, UX đông nút", "Breakpoint 768"],
    ["B3", "B. Bảng Ops", "BOOK", "Thêm booking (phím N)", "Tạo lô trống ngày/kho đang chọn", "NV kho / VP", "★★★★★", 5, "Có", "Booking tay khi chưa có Sheet", "", "blankShipmentDraft"],
    ["B4", "B. Bảng Ops", "AWB", "Sửa AWB 11 số", "Format XXX-XXXX XXXX; unique khi đủ 11 số", "Tất cả", "★★★★★", 5, "Có", "Chuẩn IATA/TCS", "AWB dở được trùng tạm — dễ rối", "Prefix 3 + number 8"],
    ["B5", "B. Bảng Ops", "FLIGHT", "Sửa chuyến / ngày bay / cutoff", "05APR, DEST, ghi chú PER…", "Tất cả", "★★★★★", 5, "Có", "Đủ field cho eSID + tem + eCargo", "", ""],
    ["B6", "B. Bảng Ops", "PCSKG", "Sửa PCS / KG", "Số kiện, trọng lượng gross", "NV cân", "★★★★★", 5, "Có", "Kích hoạt auto-status RECEIVED", "", "Checklist portal"],
    ["B7", "B. Bảng Ops", "CUST", "Gắn khách (Customer Code)", "Picker / gợi ý danh bạ 2–5 chữ A–Z", "Tất cả", "★★★★★", 5, "Có", "Nguồn shipper–CNEE–hàng", "Phụ thuộc danh bạ đầy đủ", ""],
    ["B8", "B. Bảng Ops", "CNEE", "Chọn / xem CNEE + copy block", "Popover chi tiết, copy", "PC kho", "★★★★", 4, "Không", "Giảm gõ tay, gửi khách", "", ""],
    ["B9", "B. Bảng Ops", "NOTE", "Ghi chú lô", "Note tự do; ghép Other Request", "Tất cả", "★★★★", 4, "Không", "Linh hoạt yêu cầu riêng", "Dễ thành chỗ chứa dữ liệu không cấu trúc", ""],
    ["B10", "B. Bảng Ops", "HAWB", "HAWB (House)", "In tem khi có house AWB", "PC kho", "★★★", 3, "Không", "Đúng khi có HAWB", "Không phải mọi lô", ""],
    ["B11", "B. Bảng Ops", "DEL", "Xóa lô (confirm)", "Xóa booking nhầm", "NV kho", "★★★★", 4, "Không", "Dọn dữ liệu", "Không có thùng rác / undo", ""],
    ["B12", "B. Bảng Ops", "WF", "Workflow trạng thái", "PENDING→…→WEIGH_SLIP/COMPLETED", "Tất cả", "★★★★★", 5, "Có", "Nhìn được bước còn thiếu", "TCS có RECEPTION; SCSC không — dễ nhầm UI", "shared/shipmentWorkflowStatus"],
    ["B13", "B. Bảng Ops", "AUTO_ST", "Auto-status AWB/PCS/DIM", "Thiếu AWB/PCS=PENDING; đủ=RECEIVED; có DIM=VOLUME_DONE", "Hệ thống", "★★★★★", 5, "Có", "NV không phải nhớ bấm sớm", "Không đè status manual sau này", ""],
    ["B14", "B. Bảng Ops", "MAN_ST", "Status thủ công", "CUSTOMS, OLA_PULL, RECEPTION, tờ cân, COMPLETED", "NV kiểm soát", "★★★★★", 5, "Có", "Khớp bước sau DIM", "", ""],
    ["B15", "B. Bảng Ops", "FILTER_ST", "Lọc trạng thái", "Thanh filter theo status", "Tất cả", "★★★★", 4, "Không", "Ca đông, chỉ xem chưa đo", "", ""],
    ["B16", "B. Bảng Ops", "SEARCH", "Tìm thông minh", "MAWB, số xe, tài xế, DEST; phím / và F", "Tất cả", "★★★★★", 5, "Không", "Tìm lô lúc cao điểm", "", ""],
    ["B17", "B. Bảng Ops", "FLT_DATE", "Lọc ngày bay", "Trong search bar", "Tất cả", "★★★★", 4, "Không", "Nhiều chuyến một ngày", "", ""],
    ["B18", "B. Bảng Ops", "WH_UI", "Chọn / lưới kho", "Filter layout 4 kho", "Tất cả", "★★★★★", 5, "Có", "Mỗi NV một kho", "", ""],
    ["B19", "B. Bảng Ops", "DATE", "Date picker phiên", "Đổi ngày ca", "Tất cả", "★★★★★", 5, "Có", "Sửa/xem ca hôm trước", "", ""],
    ["B20", "B. Bảng Ops", "KPI", "KPI mini lô / kiện / kg", "Header mobile", "NV kho", "★★★", 3, "Không", "Nhìn nhanh sản lượng", "Không thay stats", ""],
    ["B21", "B. Bảng Ops", "COLLAPSE", "Thu gọn kho 0 lô (mobile)", "Ít cuộn", "Phone", "★★", 2, "Không", "Tiện", "Không bắt buộc", ""],
    ["B22", "B. Bảng Ops", "COPY_UI", "Copy / phóng to chữ", "Hover magnify, copy popover", "PC", "★★★", 3, "Không", "Đọc AWB trên màn nhỏ", "", ""],
    ["B23", "B. Bảng Ops", "PRINT_SNAP", "Snapshot *Print trên lô", "Giữ shipper/CNEE/hàng lúc in", "Hệ thống", "★★★★★", 5, "Có", "Đổi danh bạ không làm sai tem đã cân", "Nhiều field trùng danh bạ", "Rất quan trọng production"],
    ["C1", "C. DIM / cước", "DIM_MOBILE", "Nhập DIM trên điện thoại", "D×R×C × số kiện", "NV cân", "★★★★★", 5, "Có", "Việc nặng nhất của ca kho", "Modal ~1500 dòng", "MobileDimKgModal"],
    ["C2", "C. DIM / cước", "DIM_PARSE", "Parse text kích thước", "Dán chuỗi kiện", "NV cân", "★★★★", 4, "Không", "Nhanh khi nhiều dòng", "Parse sai nếu format lạ", ""],
    ["C3", "C. DIM / cước", "DIVISOR", "Divisor 6000 / 5000", "Hệ số volumetric", "NV cân / rule", "★★★★★", 5, "Có", "Đúng IATA/TCS", "Phải chọn đúng hãng", "volumetricDim.ts"],
    ["C4", "C. DIM / cước", "SCSC_CW", "Chargeable SCSC theo hãng", "Làm tròn dòng/tổng (22+ hãng, VD CX)", "Kho SCSC", "★★★★★", 5, "Có (SCSC)", "Tránh tranh chấp cân", "Rule phức tạp, phải cập nhật khi hãng đổi", "scscChargeableWeight"],
    ["C5", "C. DIM / cước", "SMART_FILL", "Random / smart fill kiện", "Ước lượng chia kiện", "NV cân", "★★★", 3, "Không", "Khi chưa đo hết", "Dễ lạm dụng → sai cước", "dimBulkFill ~1579 LOC"],
    ["C6", "C. DIM / cước", "DIM_TPL", "Template / preset DIM", "Lưu mẫu kiện hay dùng", "NV cân", "★★★", 3, "Không", "KH gửi hàng lặp size", "", ""],
    ["C7", "C. DIM / cước", "LIMIT", "Cảnh báo limit hãng SCSC", "Oversize / quá kg", "NV cân", "★★★★", 4, "Không", "Chặn trước khi khai", "", ""],
    ["D1", "D. Cổng TCS eSID", "TCS_LOGIN", "Đăng nhập portal TCS", "Ext OCR CAPTCHA hoặc agent", "PC kho", "★★★★★", 5, "Có (TCS)", "ĐN được mỗi ngày", "Phụ thuộc DOM/CAPTCHA TCS", ""],
    ["D2", "D. Cổng TCS eSID", "SCAN", "Quét tiếp nhận HT", "Đối soát AWB portal → RECEPTION_COMPLETED", "PC / phone+agent", "★★★★★", 5, "Có (TCS)", "Biết lô nào HT thật", "Phone chỉ agent, không Ext", "Không đè WEIGH_SLIP"],
    ["D3", "D. Cổng TCS eSID", "FILL", "Điền eSID từ danh bạ", "Payload registrant+agent+party+AWB; không auto submit", "PC + Ext", "★★★★★", 5, "Có (TCS)", "Tiết kiệm 5–15 phút/lô", "Ext gần fork 2 bản; hook ~1255 dòng", "Chỉ PC, ẩn trên mobile"],
    ["D4", "D. Cổng TCS eSID", "SUBMIT", "HOÀN TẤT eSID (người bấm)", "Submit sau khi kiểm form", "PC kho", "★★★★★", 5, "Có (TCS)", "An toàn — không phiếu lỗi hàng loạt", "Vẫn phụ thuộc NV kiểm", "Cấm auto-submit là đúng"],
    ["D5", "D. Cổng TCS eSID", "PDF_ESID", "Tải PDF eSID", "Lấy phiếu đã khai", "PC / agent", "★★★★", 4, "Không", "Hồ sơ / gửi khách", "Hay cần agent hơn Ext", ""],
    ["D6", "D. Cổng TCS eSID", "REGISTRANT", "Hồ sơ Người khai CCCD/SĐT", "Dùng chung mọi máy", "VP / PC", "★★★★★", 5, "Có (TCS)", "Portal bắt buộc", "PII trên server", ""],
    ["D7", "D. Cổng TCS eSID", "AGENT", "Hồ sơ Agent eSID", "Agent cố định", "VP", "★★★★★", 5, "Có (TCS)", "Thiếu là Điền bị chặn", "", ""],
    ["D8", "D. Cổng TCS eSID", "QFILL", "Quick Fill từ danh bạ", "Điền eSID không cần lô Ops", "VP", "★★★", 3, "Không", "Việc lẻ", "Lệch so với vòng ca", ""],
    ["D9", "D. Cổng TCS eSID", "POLICY", "Policy Ext / Visual / Playwright", "Chọn kênh thực thi", "PC", "★★★", 3, "Không", "Linh hoạt thiết bị", "3 kênh chồng, khó hiểu", "portalExecutorPolicy"],
    ["D10", "D. Cổng TCS eSID", "DUAL", "Dual account portal", "Hub :8765 ≠ TCS :8766; Ext profile tách", "IT / máy kho", "★★★★★", 5, "Có (TCS)", "Đúng 2 tài khoản kho", "Vận hành phức tạp, dễ ĐN nhầm", "X-Portal-Warehouse"],
    ["D11", "D. Cổng TCS eSID", "EXT_DL", "Tải gói Chrome Extension", "Zip TECS-TCS / TCS / SCSC", "IT / PC mới", "★★★★", 4, "Không", "Cài máy mới nhanh", "3 zip, OCR nhân đôi", ""],
    ["D12", "D. Cổng TCS eSID", "PW_AGENT", "Agent Playwright", "Scan/PDF/fallback headless hoặc headed", "Máy kho / Railway", "★★★★", 4, "Không*", "Phone không Ext vẫn Quét/PDF", "Page object 2–3k dòng; nặng bảo trì", "*Cần nếu không ngồi Chrome"],
    ["D13", "D. Cổng TCS eSID", "JOBQ", "Portal job queue phone→PC", "Job remote worker", "Phone", "★★", 2, "Không", "Ý tưởng đa thiết bị", "Một phần deprecate, chồng kênh", "Nên cắt khi viết mới"],
    ["E1", "E. eCargo SCSC", "VCT", "Đăng ký xe VCT", "Điền form ecargo.scsc.vn", "Kho SCSC", "★★★★★", 5, "Có (SCSC)", "Xe không vào kho nếu thiếu phiếu", "Modal ~1332 dòng", "Chỉ warehouse SCSC"],
    ["E2", "E. eCargo SCSC", "EPRO", "Hồ sơ đại lý / PIC / email OTP", "Profile eCargo", "VP / SCSC", "★★★★★", 5, "Có (SCSC)", "OTP phải trùng mailbox", "", ""],
    ["E3", "E. eCargo SCSC", "IMAP", "Đọc OTP Gmail (IMAP)", "Server poll mã + link «đây»", "Hệ thống", "★★★★★", 5, "Có (SCSC)", "Hoàn tất VCT", "Phụ thuộc App Password Gmail", "imapflow"],
    ["E4", "E. eCargo SCSC", "TURN", "Turnstile / xác thực cổng", "Ext điền + confirm", "PC SCSC", "★★★★★", 5, "Có (SCSC)", "Bước cổng bắt buộc", "Dễ gãy khi SCSC đổi widget", ""],
    ["E5", "E. eCargo SCSC", "QR", "Lưu QR / mã phiếu VCT", "Kết quả đăng ký", "Tài xế / NV", "★★★★", 4, "Không", "Xuất trình cửa kho", "", ""],
    ["E6", "E. eCargo SCSC", "VEH", "Xe / loại GT từ danh bạ", "savedVehicles", "SCSC", "★★★★", 4, "Không", "Biển số lặp lại", "Biển OTO 7–9 ký tự", "≥90 phút trước giờ vào"],
    ["E7", "E. eCargo SCSC", "GATE", "Chặn VCT trên TECS-SCSC", "Chỉ SCSC trực tiếp", "Hệ thống", "★★★★★", 5, "Có", "Tránh đăng ký nhầm cổng", "", ""],
    ["F1", "F. In & xuất", "LABEL", "In nhãn nhiệt", "100×80 / 100×50; số tem nhập tay; Origin SGN", "PC kho", "★★★★★", 5, "Có", "Dán kiện XP-470B", "Số tem không auto theo PCS", "Đúng máy đang dùng"],
    ["F2", "F. In & xuất", "AIRLINE_LB", "Cấu hình tên hãng trên tem", "Override airline label", "VP", "★★★", 3, "Không", "Tem đúng thương hiệu", "", ""],
    ["F3", "F. In & xuất", "DIM_SCSC_P", "In LIST DIM SCSC", "Phiếu DIM theo lô", "Kho SCSC", "★★★★★", 5, "Có (SCSC)", "Nộp kho SCSC", "", ""],
    ["F4", "F. In & xuất", "DIM_SCSC_X", "Excel LIST DIM SCSC", "File theo lô hoặc cả ngày", "VP / SCSC", "★★★★", 4, "Không", "Gửi kế toán/kho", "", ""],
    ["F5", "F. In & xuất", "DIM_TCS_P", "In LIST DIM TCS", "Đính kèm volume TCS", "Kho TCS", "★★★★", 4, "Không", "Hồ sơ TCS", "", ""],
    ["F6", "F. In & xuất", "DIM_TCS_X", "Excel ATTACHED DIM TCS", "File đính kèm", "Kho TCS", "★★★★", 4, "Không", "Cùng F5 dạng Excel", "Hai định dạng hơi trùng", ""],
    ["F7", "F. In & xuất", "DIM_PDF", "PDF DIM TCS QF/ED/49", "Form hãng", "Một số hãng", "★★★★", 4, "Không", "Hãng bắt form này", "Không mọi chuyến", ""],
    ["F8", "F. In & xuất", "CSD", "In CSD FD / TH", "Thai AirAsia / Thai Airways + transit", "PC", "★★★", 3, "Không", "Đúng vài chuyến", "Hẹp; chỉ FD/TH", ""],
    ["F9", "F. In & xuất", "XLS_DAY", "Xuất Excel ngày / khoảng ngày", "Báo cáo ca", "VP", "★★★★", 4, "Không", "Đối soát văn phòng", "", ""],
    ["F10", "F. In & xuất", "IMG_RPT", "Copy ảnh báo cáo 4 team", "Vantage / Tecs / TCS / SCSC", "Quản lý / Zalo", "★★★★", 4, "Không", "Gửi chat rất nhanh, đúng tách kho", "Phụ thuộc clipboard", "OpsTeam ≠ Family"],
    ["F11", "F. In & xuất", "ESID_XLS", "Excel dry-run eSID", "Xem payload trước điền", "IT / QC", "★★★", 3, "Không", "Debug", "User thường không dùng", ""],
    ["G1", "G. Google Sheet", "SHEET", "Import BOOK HẰNG NGÀY", "Tab DDMONTHYYYY → lô", "Văn phòng", "★★★★★", 5, "Có", "Đầu ca không gõ từng booking", "Lệch cột Sheet thì đau", ""],
    ["G2", "G. Google Sheet", "RECON", "Reconcile AWB / fingerprint", "Không trùng phiên; ghép lô trống AWB", "Hệ thống", "★★★★★", 5, "Có", "Chống nhân đôi", "Fingerprint phức tạp", "sheetRowReconcile"],
    ["G3", "G. Google Sheet", "AI_SHEET", "AI giải thích dòng Sheet", "Gemini optional", "VP", "★★", 2, "Không", "Khi lệch cột", "Cần GEMINI_API_KEY", ""],
    ["H1", "H. Danh bạ KH", "CRUD_KH", "CRUD khách + mã 2–5 A–Z", "Khóa đồng bộ", "VP", "★★★★★", 5, "Có", "Một mã = một KH mọi máy", "", ""],
    ["H2", "H. Danh bạ KH", "SHIPPER", "Shipper đa chi nhánh", "Tên, ĐC, SĐT, MST", "VP / PC", "★★★★★", 5, "Có", "eSID / tem", "De-dup theo tên+ĐC phải cẩn", ""],
    ["H3", "H. Danh bạ KH", "CNEE_SV", "CNEE lưu sẵn", "Nhiều người nhận", "VP", "★★★★★", 5, "Có", "Cùng KH nhiều DEST", "", ""],
    ["H4", "H. Danh bạ KH", "GOODS", "Tên hàng lưu sẵn", "Goods description", "VP", "★★★★", 4, "Không", "Mô tả eSID", "", ""],
    ["H5", "H. Danh bạ KH", "VEH_SV", "Xe / loại GT lưu sẵn", "eCargo", "SCSC", "★★★★", 4, "Không", "Biển số lặp", "Chỉ nhánh SCSC", ""],
    ["H6", "H. Danh bạ KH", "DEFAULTS", "Default shipper/CNEE/goods", "Ưu tiên khi resolve eSID", "Hệ thống", "★★★★★", 5, "Có", "Không chọn tay mỗi lô", "", "resolveShipmentForEsidDeclare"],
    ["H7", "H. Danh bạ KH", "XLS_IMP", "Import/export Excel hồ sơ ~21 cột", "Onboard hàng loạt", "VP", "★★★★", 4, "Không", "Nạp KH nhanh", "Merge không được xóa saved*", ""],
    ["H8", "H. Danh bạ KH", "XLS_TPL", "Tải template Excel danh bạ", "Mẫu cột", "VP", "★★★", 3, "Không", "Hướng dẫn nhập", "", ""],
    ["H9", "H. Danh bạ KH", "VALID", "Validate mã / trùng / độ dài", "Chặn lưu sai", "Hệ thống", "★★★★★", 5, "Có", "Toàn vẹn danh bạ", "", ""],
    ["H10", "H. Danh bạ KH", "RATE", "Loại KH / đơn giá mặc định", "Account fields", "Kế toán", "★★", 2, "Không", "Phụ billing", "Không phải ca kho", ""],
    ["H11", "H. Danh bạ KH", "LEGACY_XLS", "Excel customs ops (legacy)", "Còn test, UI không dùng", "—", "★", 1, "Không", "", "Dead path", "Nên xóa khi viết mới"],
    ["I1", "I. Thống kê", "KPI_ST", "KPI lô / kg / DIM / chargeable", "Theo kỳ", "Quản lý", "★★★★", 4, "Không", "Nhìn sản lượng", "Không chặn ca", "#/stats"],
    ["I2", "I. Thống kê", "PERIOD", "Lọc hôm nay / ngày / tháng / năm / khoảng", "Period picker", "Quản lý", "★★★★", 4, "Không", "Báo cáo tuần/tháng", "", ""],
    ["I3", "I. Thống kê", "CHART", "Biểu đồ kho / DEST / xu hướng", "Recharts", "Quản lý", "★★★", 3, "Không", "Pattern", "Có thể thay bằng Excel lúc đầu", ""],
    ["I4", "I. Thống kê", "LOT_TAB", "Chi tiết lô trong kỳ", "Tab lots", "VP", "★★★", 3, "Không", "Đối soát", "", ""],
    ["I5", "I. Thống kê", "XLS_ST", "Xuất Excel stats", "File", "VP", "★★★", 3, "Không", "Gửi sếp", "", ""],
    ["I6", "I. Thống kê", "AI_EOD", "Tóm tắt cuối ngày AI", "Gemini", "Quản lý", "★★", 2, "Không", "Nhanh", "Optional", ""],
    ["J1", "J. AI Gemini", "AI_OPS", "Trợ lý AI Ops", "Parse booking/DIM; draft→Confirm", "VP / thử nghiệm", "★★", 2, "Không", "Giảm gõ", "Không được tự submit portal; cần key", "Feature flag"],
    ["J2", "J. AI Gemini", "AI_RPT", "Báo cáo đề xuất nâng cấp", "Phân tích telemetry UI", "Dev", "★", 1, "Không", "Cho đội kỹ thuật", "Không giúp NV kho", ""],
    ["J3", "J. AI Gemini", "AI_TEL", "Telemetry sự kiện AI", "ops_ai_events", "Dev", "★", 1, "Không", "Cải tiến sản phẩm", "Overhead", ""],
    ["K1", "K. Hạ tầng", "BACKUP", "Backup / restore Postgres", "Scripts", "IT", "★★★★", 4, "Không", "An toàn dữ liệu", "Phải nhớ chạy", ""],
    ["K2", "K. Hạ tầng", "DEPLOY", "Deploy Railway + healthcheck", "deploy:ship, /api/health", "IT", "★★★★★", 5, "Có", "App lên được", "Nhiều script lịch sử", ""],
    ["K3", "K. Hạ tầng", "CATALOG", "Catalog hãng / sân bay", "Master data", "Hệ thống", "★★★★", 4, "Có", "Tem, DEST, rule SCSC", "", ""],
    ["K4", "K. Hạ tầng", "OCR", "OCR CAPTCHA trong Ext (ONNX)", "ĐN TCS offline model", "PC kho", "★★★★★", 5, "Có (PC Ext)", "ĐN không bắt buộc agent", "Binary OCR / Docker fetch phức tạp", ""],
    ["K5", "K. Hạ tầng", "PYSIDE", "Desktop PySide6 sidecar", "UI Python agent", "—", "★", 1, "Không", "", "Ops đi web; đường phụ", "Nên không đưa vào sản phẩm mới"],
    ["K6", "K. Hạ tầng", "E2E", "E2E Playwright / a11y / unit test", "Vitest + e2e L1/L2", "Dev", "★★★", 3, "Không", "Khóa regression domain", "Ext/page object thiếu coverage", "~10k LOC test vs ~78k prod"],
]

STRENGTHS = [
    [1, "Nghiệp vụ", "Khóa được vòng ca thật", "Booking → DIM → eSID/eCargo → in → đóng ca; không phải CRUD demo.", "Rất cao", "Giữ nguyên khi viết mới"],
    [2, "Nghiệp vụ", "4 kho + dual account portal", "Tách TECS-TCS / TCS / SCSC đúng thực tế 2 user TCS khác nhau.", "Rất cao", "Giữ invariant"],
    [3, "Nghiệp vụ", "Không auto-submit cổng ngoài", "Điền form, người kiểm tra rồi HOÀN TẤT — tránh phiếu eSID lỗi hàng loạt.", "Rất cao", "Giữ"],
    [4, "Nghiệp vụ", "Auto-status AWB / PCS / DIM", "NV cân không phải nhớ bấm trạng thái sớm.", "Cao", "Giữ + test"],
    [5, "Nghiệp vụ", "Rule DIM / chargeable SCSC theo hãng", "Đúng tranh chấp cân với kho/hãng.", "Rất cao", "Port sang packages/domain"],
    [6, "Nghiệp vụ", "Snapshot *Print trên lô", "Đổi danh bạ không làm sai tem/phiếu đã cân.", "Cao", "Giữ schema"],
    [7, "Đa thiết bị", "Phone DIM + PC portal cùng một nguồn dữ liệu", "NV cân và NV kiểm soát làm song song.", "Rất cao", "Realtime + PWA"],
    [8, "Đa thiết bị", "Offline queue", "Wifi kho chết không mất thao tác (có trần 500).", "Cao", "IndexedDB + idempotency"],
    [9, "Tích hợp", "Chrome Ext điền nhìn thấy", "Cookie/CAPTCHA/session Chrome thật — ổn định hơn headless thuần.", "Rất cao", "Gộp 2 Ext eSID"],
    [10, "Tích hợp", "IMAP OTP eCargo", "Khép vòng VCT — ít app nội bộ làm được.", "Rất cao", "Giữ API Railway"],
    [11, "Tích hợp", "Google Sheet BOOK HẰNG NGÀY + reconcile", "Đầu ca văn phòng; chống trùng AWB/fingerprint.", "Cao", "Giữ parser"],
    [12, "Tích hợp", "In tem nhiệt đúng khổ máy", "100×80 / 100×50, số bản tay, XP-470B.", "Cao", "Giữ"],
    [13, "Dữ liệu", "Postgres bắt buộc trên Railway", "Redeploy không mất ca (đã học từ mất data container).", "Rất cao", "Supabase/Postgres"],
    [14, "Dữ liệu", "Danh bạ đa chi nhánh + Customer Code", "Một mã KH map shipper/CNEE/goods/xe.", "Cao", "Schema quan hệ"],
    [15, "Báo cáo", "Ảnh báo cáo 4 team (Vantage/Tecs/TCS/SCSC)", "Gửi Zalo đúng tách kho, không gộp nhầm hub.", "Cao", "Giữ rule OpsTeam"],
    [16, "Chất lượng", "Unit test domain khá dày", "AWB, Sheet reconcile, DIM, validation có test.", "Trung bình–cao", "Giữ khi port domain"],
    [17, "UX", "Desktop + mobile cùng sản phẩm", "Không cần app native riêng lúc này.", "Cao", "PWA hóa"],
    [18, "An toàn nghiệp vụ", "Checklist trước Fill/Register", "Cảnh báo thiếu chuyến/DEST/PCS/KG/mã KH.", "Cao", "Giữ"],
]

WEAKNESSES = [
    [1, "Kiến trúc", "Cao", "Code nặng hơn bề mặt UI", "~88k LOC / 3 màn hash. God-file: DIM modal, Ext, Playwright, useTcsPortalActions.", "Gộp Ext, tách domain, cắt legacy", "Viết mới / refactor"],
    [2, "Automation", "Cao", "3 kênh portal chồng", "Ext + Playwright agent + portal-jobs (một phần deprecate).", "Một kênh chính (Ext) + một phụ (agent)", "Viết mới"],
    [3, "Automation", "Cao", "Near-fork 2 Chrome Ext eSID", "~9k dòng content+background gần trùng; fix DOM dễ lệch bản.", "1 package + config kho", "Ưu tiên P0 viết mới"],
    [4, "Automation", "Cao", "Page object Playwright cực lớn", "esid_declare_page ~2909, esid_page ~2113 — bảo trì DOM đắt.", "Adapter mỏng + locator JSON versioned", "Viết mới"],
    [5, "Dữ liệu", "Trung bình", "Dual-write blob JSONB + bảng quan hệ", "An toàn migrate nhưng phức tạp, dễ lệch schema.", "Một schema quan hệ (Supabase)", "Viết mới"],
    [6, "Dữ liệu", "Trung bình", "Mutation bus tự viết thay REST/resource", "Khó onboarding; mọi entity đi một ống.", "API rõ + Realtime theo bảng", "Viết mới"],
    [7, "Bảo mật", "Trung bình", "Auth token dùng chung", "Chưa role NV cân / kiểm soát / admin.", "Supabase Auth + RLS theo role", "P1"],
    [8, "Bảo mật", "Trung bình", "PII (CCCD, mật khẩu portal) trên máy/Ext", "Cần kiểm soát nhớ mật, OCR, log.", "Secret manager; không log PII", "Song song"],
    [9, "Phụ thuộc ngoài", "Cao", "Lệ thuộc DOM TCS / eCargo / CAPTCHA", "Cổng đổi là Ext/agent gãy — rủi ro lớn nhất của sản phẩm.", "Contract adapter; test locator; human confirm", "Không tránh được — giảm phạm vi"],
    [10, "Phụ thuộc ngoài", "Trung bình", "Gmail IMAP App Password cho OTP", "Đứt mail / đổi 2FA là kẹt VCT.", "Mailbox riêng + cảnh báo health IMAP", "Giữ + monitor"],
    [11, "Chất lượng", "Cao", "Test không cân phần rủi ro cao", "Utils/server khá; Ext + page object gần như thủ công.", "CDP verify + contract test payload", "P1"],
    [12, "UX", "Trung bình", "Ops nhồi quá nhiều nút", "AirCargoTracking orchestration; mobile sticky đông.", "Tách màn theo vai trò (cân / kiểm soát)", "UX viết mới"],
    [13, "UX", "Thấp", "Smart fill DIM dễ lạm dụng", "Ước lượng kiện có thể sai cước.", "Nhãn «ước lượng» + cấm xuất chính thức", "Rule"],
    [14, "Sản phẩm", "Thấp–TB", "AI Gemini không nằm critical path", "~subsystem đầy đủ nhưng NV kho ít dùng.", "Feature flag / plugin, không nhồi ngày 1", "Cắt khỏi lõi"],
    [15, "Sản phẩm", "Thấp", "Legacy còn sót", "customerCustomsOpsExcel, portalRemoteJobs, PySide6, scripts deprecate.", "Xóa khỏi sản phẩm mới", "Dọn"],
    [16, "Vận hành", "Trung bình", "Cài Ext + dual Chrome profile", "Máy mới / NV mới onboarding nặng.", "Installer agent + 1 Ext; hướng dẫn 1 trang", "Máy kho"],
    [17, "Vận hành", "Trung bình", "Nhiều script deploy lịch sử", "Khó biết lệnh nào còn sống.", "1 pipeline Railway + 1 docs", "Dọn"],
    [18, "Hiệu năng", "Thấp–TB", "Chunk UI từng rất lớn (đã code-split)", "ExcelJS/PDF vẫn nặng lúc export.", "Lazy đã có; giữ", "Monitor"],
]

PRIORITY = [
    ["P0 — Phải có để thay TECSOPS", "A2,A3,A4,A5,A7,B1–B7,B12–B14,B18,B19,B23,C1,C3,C4,D1–D4,D6,D7,D10,E1–E4,E7,F1,F3,G1,G2,H1–H3,H6,H9,K2,K4", "Không có thì dừng ca hoặc sai nghiệp vụ"],
    ["P1 — Nên có ngay sau lõi", "A1,A6,B8,B9,B11,B15–B17,C2,C7,D5,D11,D12,E5,E6,F4–F7,F9,F10,H4,H5,H7,I1,I2,K1,K3", "Tăng tốc ca / văn phòng"],
    ["P2 — Làm sau", "B10,B20,B22,C5,C6,D8,D9,F2,F8,F11,H8,I3–I5", "Hữu ích hẹp hoặc thay được Excel"],
    ["P3 — Cắt / optional", "B21,D13,G3,H10,H11,I6,J1–J3,K5", "Không chặn ca; gây nặng code"],
]

USERS = [
    ["NV cân kiện", "Phone", "Ngày phiên, booking/sửa PCS-KG, DIM, tìm lô, status, sync/offline", "A2–A6, B2–B7, B12–B19, C1–C4, C7", "Không Điền eSID (policy mobile = agent Quét/PDF)"],
    ["NV kiểm soát kho TCS", "PC + Chrome Ext TECS-TCS hoặc TCS", "ĐN portal, Điền, HOÀN TẤT, Quét HT, PDF, in tem, LIST DIM", "D1–D7, D10–D12, F1, F5–F7", "Profile Chrome tách theo kho"],
    ["NV kho SCSC", "PC + Ext SCSC", "DIM rule hãng, LIST DIM SCSC, eCargo VCT + OTP + QR", "C4, E1–E7, F3, F4", "Không eSID TCS; TECS-SCSC không VCT"],
    ["Văn phòng / CS", "Laptop", "Import Sheet, danh bạ Excel, Người khai/Agent, báo cáo ngày", "G1–G2, H1–H9, D6–D7, F9–F10", "Không thay luật kho"],
    ["Quản lý", "Laptop", "Stats KPI, ảnh 4 team, Excel stats", "I1–I5, F10", "Không cần portal"],
    ["IT / Dev", "Máy kho + Railway", "Cài Ext, dual agent, backup, deploy, OCR", "K1–K4, D10–D12", "Onboarding còn nặng"],
]


def build() -> None:
    wb = Workbook()

    # ----- 0. Huong dan -----
    ws0 = wb.active
    ws0.title = "0. Huong_dan"
    ws0.sheet_properties.tabColor = NAVY
    ws0.merge_cells("B2:G2")
    ws0["B2"] = "TECSOPS — Danh mục chức năng & điểm mạnh / điểm yếu"
    ws0["B2"].font = FONT_TITLE
    ws0.merge_cells("B3:G3")
    ws0["B3"] = "Hệ thống Quản lý Vận hành Hàng không & Hải quan TCS / SCSC  ·  File sinh từ audit mã nguồn  ·  2026-08-14"
    ws0["B3"].font = FONT_SMALL

    intro = [
        (5, "Mục đích file", "Liệt kê chức năng đang có, chấm công dụng vận hành ngày, và nêu bật điểm mạnh / yếu để quyết định giữ–cắt–viết mới."),
        (6, "Cách đọc", "Sheet 1 tóm tắt. Sheet 2 lọc theo nhóm / điểm / bắt buộc ca. Sheet 3–4 là SWOT. Sheet 5 ưu tiên viết siêu app. Sheet 6 theo vai trò."),
        (7, "Thang ★", "★★★★★ dừng ca nếu thiếu  ·  ★★★★ dùng mỗi ngày  ·  ★★★ hữu ích  ·  ★★ phụ trợ  ·  ★ optional / legacy"),
        (8, "Cột Điểm", "1–5 trùng số sao. Màu: xanh đậm = lõi, xanh dương = nên có, vàng = hỗ trợ, cam = phụ, đỏ = cắt."),
        (9, "Phạm vi", "Ops #/  ·  Customers #/customers  ·  Stats #/stats  ·  Chrome Ext ×3  ·  Playwright agent  ·  Server Railway/Postgres"),
        (10, "Không gồm", "Docs/memory nội bộ agent, script one-off, file test trừ khi là chức năng user."),
    ]
    ws0["B5"].font = FONT_H2
    for row, title, text in intro:
        ws0.cell(row, 2, title).font = FONT_BOLD
        ws0.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        ws0.cell(row, 3, text).font = FONT_CELL
        ws0.cell(row, 3).alignment = WRAP

    legend = [
        (12, "Màu", "Ý nghĩa"),
        (13, "★★★★★ / 5", "Lõi production — phải có nếu muốn dùng thật như hiện tại"),
        (14, "★★★★ / 4", "Tăng tốc ca hoặc văn phòng — nên có ngay sau lõi"),
        (15, "★★★ / 3", "Hữu ích hẹp (HAWB, CSD, chart…)"),
        (16, "★★ / 2", "Tiện ích / AI / UX phụ"),
        (17, "★ / 1", "Legacy hoặc không phải chức năng NV kho"),
    ]
    ws0.cell(12, 2, "Màu").font = FONT_HEAD
    ws0.cell(12, 2).fill = fill(NAVY)
    ws0.cell(12, 3, "Ý nghĩa").font = FONT_HEAD
    ws0.cell(12, 3).fill = fill(NAVY)
    paints = [GREEN_BG, BLUE_BG, YELLOW_BG, ORANGE_BG, RED_BG]
    fonts = [GREEN, BLUE, YELLOW, ORANGE, RED]
    for i, (row, a, b) in enumerate(legend[1:], 0):
        ws0.cell(row, 2, a).fill = fill(paints[i])
        ws0.cell(row, 2).font = Font(name="Calibri", size=10, bold=True, color=fonts[i])
        ws0.cell(row, 3, b).font = FONT_CELL
        ws0.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)

    ws0.merge_cells("B19:G19")
    ws0["B19"] = "Kết luận nhanh: ~73 chức năng liệt kê; ~35 cái điểm 5 là xương sống ca thật. Điểm mạnh là nghiệp vụ cổng TCS/SCSC + sync đa máy. Điểm yếu là độ nặng code, fork Ext, và 3 kênh automation."
    ws0["B19"].font = Font(name="Calibri", size=11, bold=True, color=NAVY2)
    ws0["B19"].alignment = WRAP
    ws0["B19"].fill = fill(GOLD_BG)

    set_widths(ws0, [3, 22, 28, 18, 18, 18, 18])
    ws0.row_dimensions[2].height = 28
    ws0.row_dimensions[19].height = 48
    for r in range(5, 11):
        ws0.row_dimensions[r].height = 32
    ws0.row_dimensions[3].height = 22
    ws0.print_options.horizontalCentered = True

    # ----- 1. Tom tat -----
    ws1 = wb.create_sheet("1. Tom_tat")
    ws1.sheet_properties.tabColor = TEAL
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "Tóm tắt đánh giá dự án TECSOPS"
    ws1["A1"].font = FONT_TITLE

    summary_heads = ["Chỉ số", "Giá trị", "Ý nghĩa", "Mạnh / yếu", "Ghi chú"]
    for c, h in enumerate(summary_heads, 1):
        cell = ws1.cell(3, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill(TEAL)
        cell.alignment = CENTER
        cell.border = THIN

    n = len(FEATURES)
    n5 = sum(1 for f in FEATURES if f[7] >= 5)
    n4 = sum(1 for f in FEATURES if f[7] == 4)
    n_core = sum(1 for f in FEATURES if f[8].startswith("Có"))
    summary_rows = [
        ["Số chức năng liệt kê", n, "Toàn bộ nút/luồng user + hạ tầng nhìn thấy được", "Trung tính", "Không đếm file test/docs"],
        ["Điểm 5 — lõi ca thật", n5, "Thiếu là dừng ca hoặc sai nghiệp vụ", "Mạnh (đủ lõi)", "Xương sống siêu app mới"],
        ["Điểm 4 — nên có", n4, "Dùng mỗi ngày, tiết kiệm thời gian", "Mạnh", "P1 viết mới"],
        ["Bắt buộc theo nhánh kho", n_core, "Có / Có(TCS) / Có(SCSC) / Có(prod)", "Mạnh", "Không phải mọi nút đều mọi kho"],
        ["Màn hình user", "3 (Ops / Khách / Stats)", "UI mỏng so với ~88k LOC", "Yếu (lệch chức năng/code)", "Nặng ở Ext + agent + utils"],
        ["LOC sản xuất (ước lượng)", "~78k (src 38k + Ext 14k + agent 13k + server 8k)", "Automation chiếm ~1/3", "Yếu bảo trì", "Viết mới kỳ vọng 40–48k"],
        ["Realtime + Postgres", "Có", "Đa máy + không mất data khi deploy", "Mạnh", "Nên chuyển Supabase Realtime"],
        ["Cổng TCS eSID", "ĐN / Điền / HOÀN TẤT / Quét / PDF", "Làm app «thật»", "Mạnh nghiệp vụ — yếu bảo trì Ext", "Human-in-the-loop"],
        ["Cổng eCargo SCSC", "VCT + IMAP OTP + QR", "Khép vòng xe vào kho", "Mạnh", "Chỉ kho SCSC"],
        ["Auth phân quyền", "Token dùng chung", "Chưa role", "Yếu", "Nên Auth + RLS"],
        ["AI Gemini", "Workbench + báo cáo + EOD", "Không critical path", "Yếu ROI so với LOC", "Plugin"],
        ["Mức rủi ro hệ thống", "Trung bình–cao", "Phụ thuộc DOM cổng ngoài + dual store + legacy", "Yếu vận hành IT", "Adapter + 1 kênh automation"],
        ["Khả năng thay thế bằng siêu app mới", "Cao nếu giữ P0", "Stack đề xuất: Supabase + React/Vite + Hono/Railway + Ext gộp + agent Tauri", "Cơ hội", "Xem sheet 5"],
    ]
    for i, row in enumerate(summary_rows):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = ws1.cell(r, c, v)
            cell.font = FONT_CELL
            cell.alignment = WRAP
            cell.border = THIN
            if i % 2:
                cell.fill = fill(ROW_ALT)
        tone = row[3]
        tone_cell = ws1.cell(r, 4)
        if "Mạnh" in tone and "Yếu" not in tone:
            tone_cell.fill = fill(GREEN_BG)
            tone_cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
        elif "Yếu" in tone:
            tone_cell.fill = fill(ORANGE_BG)
            tone_cell.font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
        elif "Cơ hội" in tone:
            tone_cell.fill = fill(PURPLE_BG)
            tone_cell.font = Font(name="Calibri", size=10, bold=True, color=PURPLE)
        else:
            tone_cell.fill = fill(BLUE_BG)

    # mini counts by group
    ws1["A19"] = "Số chức năng theo nhóm"
    ws1["A19"].font = FONT_H2
    ws1["A20"] = "Nhóm"
    ws1["B20"] = "Số chức năng"
    ws1["C20"] = "Điểm TB"
    ws1["D20"] = "Số điểm 5"
    for c in range(1, 5):
        ws1.cell(20, c).font = FONT_HEAD
        ws1.cell(20, c).fill = fill(NAVY)
        ws1.cell(20, c).alignment = CENTER

    groups = []
    seen = []
    for f in FEATURES:
        g = f[1]
        if g not in seen:
            seen.append(g)
            items = [x for x in FEATURES if x[1] == g]
            avg = round(sum(x[7] for x in items) / len(items), 2)
            c5 = sum(1 for x in items if x[7] >= 5)
            groups.append((g, len(items), avg, c5))
    for i, (g, cnt, avg, c5) in enumerate(groups):
        r = 21 + i
        ws1.cell(r, 1, g).font = FONT_CELL
        ws1.cell(r, 2, cnt).alignment = CENTER
        ws1.cell(r, 3, avg).alignment = CENTER
        ws1.cell(r, 4, c5).alignment = CENTER
        for c in range(1, 5):
            ws1.cell(r, c).border = THIN
            if i % 2:
                ws1.cell(r, c).fill = fill(ROW_ALT)
        paint_score(ws1.cell(r, 3), avg)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Số chức năng theo nhóm"
    chart.y_axis.title = "Số lượng"
    data = Reference(ws1, min_col=2, min_row=20, max_row=20 + len(groups))
    cats = Reference(ws1, min_col=1, min_row=21, max_row=20 + len(groups))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.scaling.min = 0
    chart.height = 8
    chart.width = 18
    ws1.add_chart(chart, "A34")

    set_widths(ws1, [36, 22, 62, 28, 42])
    for r in range(4, 17):
        ws1.row_dimensions[r].height = 36
    ws1.freeze_panes = "A4"
    ws1.auto_filter.ref = "A3:E16"
    ws1.sheet_view.showGridLines = False
    ws1.page_setup.orientation = "landscape"
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0

    # ----- 2. Chuc nang -----
    ws2 = wb.create_sheet("2. Danh_muc_chuc_nang")
    ws2.sheet_properties.tabColor = "2563EB"
    ws2.merge_cells("A1:L1")
    ws2["A1"] = "Danh mục chi tiết chức năng — lọc theo nhóm, điểm, bắt buộc ca"
    ws2["A1"].font = FONT_TITLE
    ws2.merge_cells("A2:L2")
    ws2["A2"] = "Mẹo: bật AutoFilter trên hàng 4. Cột «Bắt buộc ca» = Có nghĩa là không thể cắt nếu muốn thay thế TECSOPS trên nhánh đó."
    ws2["A2"].font = FONT_SMALL

    heads = [
        "ID",
        "Nhóm",
        "Mã",
        "Tên chức năng",
        "Việc làm / mô tả",
        "Công dụng",
        "Điểm (1–5)",
        "Ai dùng",
        "Bắt buộc ca",
        "Điểm mạnh của chức năng",
        "Điểm yếu / rủi ro",
        "Ghi chú kỹ thuật",
    ]
    for c, h in enumerate(heads, 1):
        cell = ws2.cell(4, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill(NAVY)
        cell.alignment = CENTER
        cell.border = THIN

    for i, f in enumerate(FEATURES):
        r = 5 + i
        # FEATURES: id, nhóm, mã, tên, việc, ai, sao, điểm, bắt buộc, mạnh, yếu, ghi chú
        # Cột sheet: id, nhóm, mã, tên, việc, sao, điểm, ai, bắt buộc, mạnh, yếu, ghi chú
        mapped = [f[0], f[1], f[2], f[3], f[4], f[6], f[7], f[5], f[8], f[9], f[10], f[11]]
        apply_row(ws2, r, mapped)
        ws2.cell(r, 1).alignment = CENTER
        ws2.cell(r, 3).alignment = CENTER
        ws2.cell(r, 8).alignment = WRAP
        ws2.row_dimensions[r].height = 42
        # highlight required
        req = f[8]
        req_cell = ws2.cell(r, 9)
        req_cell.alignment = CENTER
        if req == "Có" or req.startswith("Có"):
            req_cell.fill = fill(GREEN_BG)
            req_cell.font = Font(name="Calibri", size=9, bold=True, color=GREEN)
        elif req.startswith("Không"):
            req_cell.fill = fill(ROW_ALT) if i % 2 else fill(WHITE)
            req_cell.font = Font(name="Calibri", size=9, color=MUTED)

    last = 4 + len(FEATURES)
    finish_filter(ws2, 4, last, 12)
    ws2.auto_filter.ref = f"A4:L{last}"
    set_widths(ws2, [8, 18, 12, 32, 48, 14, 12, 20, 14, 38, 40, 32])
    ws2.auto_filter.ref = f"A4:L{last}"

    # ----- 3. Diem manh -----
    ws3 = wb.create_sheet("3. Diem_manh")
    ws3.sheet_properties.tabColor = "16A34A"
    ws3.merge_cells("A1:F1")
    ws3["A1"] = "Điểm mạnh — những gì làm TECSOPS dùng được thật mỗi ngày"
    ws3["A1"].font = FONT_TITLE
    ws3.merge_cells("A2:F2")
    ws3["A2"] = "Giữ các điểm này khi viết siêu app mới. Đây là lợi thế cạnh tranh so với spreadsheet / app CRUD."
    ws3["A2"].font = FONT_SMALL

    h3 = ["#", "Trụ", "Điểm mạnh", "Giải thích", "Mức ảnh hưởng", "Hướng xử lý khi viết mới"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(4, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill("166534")
        cell.alignment = CENTER
        cell.border = THIN
    for i, row in enumerate(STRENGTHS):
        r = 5 + i
        for c, v in enumerate(row, 1):
            cell = ws3.cell(r, c, v)
            cell.font = FONT_CELL
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill(GREEN_BG) if i % 2 == 0 else fill("F0FDF4")
        ws3.cell(r, 1).alignment = CENTER
        ws3.cell(r, 5).alignment = CENTER
        if "Rất cao" in str(row[4]):
            ws3.cell(r, 5).font = Font(name="Calibri", size=10, bold=True, color=GREEN)
        ws3.row_dimensions[r].height = 40
    finish_filter(ws3, 4, 4 + len(STRENGTHS), 6)
    ws3.auto_filter.ref = f"A4:F{4 + len(STRENGTHS)}"
    set_widths(ws3, [6, 16, 42, 62, 18, 36])

    # ----- 4. Diem yeu -----
    ws4 = wb.create_sheet("4. Diem_yeu")
    ws4.sheet_properties.tabColor = "DC2626"
    ws4.merge_cells("A1:G1")
    ws4["A1"] = "Điểm yếu / rủi ro — nơi dự án nặng, dễ gãy, hoặc lệch chức năng"
    ws4["A1"].font = FONT_TITLE
    ws4.merge_cells("A2:G2")
    ws4["A2"] = "Đây là lý do viết mới có thể giảm 40–55% LOC mà vẫn giữ năng lực ops — nếu cắt đúng chỗ, không cắt P0."
    ws4["A2"].font = FONT_SMALL

    h4 = ["#", "Trụ", "Mức", "Điểm yếu", "Hiện trạng", "Hướng xử lý", "Khi nào làm"]
    for c, h in enumerate(h4, 1):
        cell = ws4.cell(4, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill("991B1B")
        cell.alignment = CENTER
        cell.border = THIN
    for i, row in enumerate(WEAKNESSES):
        r = 5 + i
        for c, v in enumerate(row, 1):
            cell = ws4.cell(r, c, v)
            cell.font = FONT_CELL
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill(RED_BG) if i % 2 == 0 else fill("FFF7ED")
        ws4.cell(r, 1).alignment = CENTER
        lvl = str(row[2])
        lvl_cell = ws4.cell(r, 3)
        lvl_cell.alignment = CENTER
        lvl_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        if lvl == "Cao":
            lvl_cell.fill = fill("DC2626")
        elif "Trung" in lvl:
            lvl_cell.fill = fill("EA580C")
        else:
            lvl_cell.fill = fill("CA8A04")
        ws4.row_dimensions[r].height = 48
    finish_filter(ws4, 4, 4 + len(WEAKNESSES), 7)
    ws4.auto_filter.ref = f"A4:G{4 + len(WEAKNESSES)}"
    set_widths(ws4, [6, 16, 14, 36, 58, 42, 22])

    # ----- 5. Uu tien -----
    ws5 = wb.create_sheet("5. Uu_tien_viet_moi")
    ws5.sheet_properties.tabColor = "7C3AED"
    ws5.merge_cells("A1:C1")
    ws5["A1"] = "Ưu tiên nếu viết siêu app mới (giữ năng lực thực tế như hiện tại)"
    ws5["A1"].font = FONT_TITLE
    h5 = ["Hạng", "ID chức năng (sheet 2)", "Nguyên tắc"]
    for c, h in enumerate(h5, 1):
        cell = ws5.cell(3, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill("6B21A8")
        cell.alignment = CENTER
        cell.border = THIN
    colors_p = [GREEN_BG, BLUE_BG, YELLOW_BG, RED_BG]
    fonts_p = [GREEN, BLUE, YELLOW, RED]
    for i, row in enumerate(PRIORITY):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = ws5.cell(r, c, v)
            cell.font = FONT_CELL
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill(colors_p[i])
        ws5.cell(r, 1).font = Font(name="Calibri", size=11, bold=True, color=fonts_p[i])
        ws5.row_dimensions[r].height = 56
    ws5.merge_cells("A9:C9")
    ws5["A9"] = (
        "Stack đề xuất khi viết mới: Supabase (Postgres + Auth + Realtime + Storage) + React/Vite/TanStack "
        "+ Hono trên Railway (IMAP, Sheets, jobs) + 1 Chrome Ext eSID + 1 Ext eCargo + agent Tauri/Playwright trên máy kho. "
        "Không copy Express/Socket/blob state. Không 3 kênh portal. Không nhồi AI ngày 1."
    )
    ws5["A9"].alignment = WRAP
    ws5["A9"].font = Font(name="Calibri", size=11, color=NAVY2)
    ws5["A9"].fill = fill(PURPLE_BG)
    ws5.row_dimensions[9].height = 64
    set_widths(ws5, [36, 88, 48])
    ws5.sheet_view.showGridLines = False

    # ----- 6. Nguoi dung -----
    ws6 = wb.create_sheet("6. Theo_nguoi_dung")
    ws6.sheet_properties.tabColor = "0369A1"
    ws6.merge_cells("A1:E1")
    ws6["A1"] = "Chức năng theo vai trò — ai dùng gì trong ca thật"
    ws6["A1"].font = FONT_TITLE
    h6 = ["Vai trò", "Thiết bị", "Việc chính trong ca", "ID chức năng liên quan", "Giới hạn / lưu ý"]
    for c, h in enumerate(h6, 1):
        cell = ws6.cell(3, c, h)
        cell.font = FONT_HEAD
        cell.fill = fill("075985")
        cell.alignment = CENTER
        cell.border = THIN
    for i, row in enumerate(USERS):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = ws6.cell(r, c, v)
            cell.font = FONT_CELL
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill(BLUE_BG) if i % 2 == 0 else fill("F0F9FF")
        ws6.row_dimensions[r].height = 52
    finish_filter(ws6, 3, 3 + len(USERS), 5)
    ws6.auto_filter.ref = f"A3:E{3 + len(USERS)}"
    set_widths(ws6, [26, 36, 62, 42, 42])

    # print / freeze niceties
    for ws in (ws0, ws1, ws2, ws3, ws4, ws5, ws6):
        ws.page_setup.orientation = "landscape"
        ws.oddHeader.left.text = "TECSOPS"
        ws.oddHeader.right.text = "Danh mục chức năng & SWOT"
        ws.oddFooter.left.text = "Nội bộ — không đưa mật khẩu / PII"
        ws.oddFooter.right.text = "Trang &P / &N"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes) features={len(FEATURES)}")


if __name__ == "__main__":
    build()
